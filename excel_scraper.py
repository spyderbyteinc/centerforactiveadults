import openpyxl
from pathlib import Path

xlsx_file = Path('yoga_new.xlsx')
wb_obj = openpyxl.load_workbook(xlsx_file) 

# Read the active sheet:
sheet = wb_obj.active

classes = []
groups = []
days = []
times = []

rows = 0
for i in range(sheet.max_column):
    count = 0
    rows = rows + 1
    for row in sheet.iter_rows():
        count = count + 1
        if(count < 13):
            val = row[i].value

            if(val == "CLASSES" or val == "Group" or val == "Day / Days" or val == "TIMES"):
                continue
            if(rows == 1):
                classes.append(val)
            elif(rows == 2):
                groups.append(val)
            elif(rows == 3):
                days.append(val)
            elif(rows == 4):
                times.append(val)


for x in range(len(classes)):
    cl = classes[x]
    gr = groups[x]
    dy = days[x]
    ti = times[x]

    print(f"INSERT INTO `classes` (`id`, `class_name`, `class_type`, `day`, `time`) VALUES (NULL, \"{cl}\", \"{gr}\", \"{dy}\", \"{ti}\");")
